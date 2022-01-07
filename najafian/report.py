""" Handles simple reports to excel files with structures from biopsy/gloms/individual files"""
import collections
import os

from os import path
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook


def generate_label(key):
    """ Create a more human readable label by splitting spaces removing underscores and cap first letters"""
    return ' '.join([label.capitalize() for label in key.replace('_', ' ').replace('file ', '').split()])


def build_report(report):
    """ Generates a global report, file report, and then an individual report

    Args:
        report ([type]): [description]

    Returns:
        [type]: [description]
    """
    g_report, f_report, i_report = [collections.OrderedDict() for _ in range(3)]
    f_t_report = {}

    # populate the individual reports
    for key in report.keys():
        if 'file' in key:
            f_t_report[key] = report[key]
        elif 'individual' in key:
            i_report[key] = report[key]
        else:
            g_report[key] = report[key]

    # populate the file report
    available_keys = list(f_t_report.keys())
    available_keys.remove('file')
    key_labels = []
    for key in available_keys:
        key_labels.append(generate_label(key))

    f_report['Name'] = ' | '.join(key_labels)
    for ind, file in enumerate(f_t_report['file']):
        f_report[file] = ''
        for kl, key in enumerate(available_keys):
            try:
                f_report[file] += '| {}: {} {}'.format(key_labels[kl], f_t_report[key][ind], '|' if key == available_keys[-1] else '')
            except IndexError:
                pass

    return g_report, f_report, i_report, f_t_report, available_keys, key_labels

def export_report(export_file, g_report, f_report, i_report, f_t_report, available_keys, key_labels):
    """ @TODO write description

    Args:
        export_file ([type]): [description]
        g_report ([type]): [description]
        f_report ([type]): [description]
        i_report ([type]): [description]
        f_t_report ([type]): [description]
        available_keys ([type]): [description]
        key_labels ([type]): [description]
    """
    # create the spreadsheet
    wb = Workbook()

    # add the global items
    ws = wb.active
    ws.title = 'Global'
    ws.append(['Item', 'Values'])
    for key in g_report.keys():
        value = g_report[key]
        items = [key]
        if isinstance(value, list):
            items.extend(value)
        else:
            items.append(str(value))
        ws.append(items)

    # add the file items
    wf = wb.create_sheet(title='File')
    header = ['Name']
    header.extend(key_labels)
    wf.append(header)
    for ind, file in enumerate(f_t_report['file']):
        items = [file]
        for kl, key in enumerate(available_keys):
            try:
                items.append(f_t_report[key][ind])
            except IndexError:
                pass
        wf.append(items)

    # add the individual items
    wi = wb.create_sheet(title='Individual')
    wi.append([generate_label(key) for key in i_report.keys()])
    data = [i_report[key] for key in i_report.keys()]
    for ind in range(len(list(data)[0])):
        wi.append([data[shift][ind] for shift in range(len(data))])

    # save the spreadsheet
    wb.save(export_file)


def compile_report(export_folder, report):
    """ Creates the report and then exports it to a specific folder """
    export_report(export_folder, *build_report(report))