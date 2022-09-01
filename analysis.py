""" General analysis file that loads the model, creates segmentation masks, and applies membrane and slit vision modules to estimate foot process width """
import argparse
import glob
import json
import logging.handlers
import os
import os.path
import os.path as path
import sys
import time

import cv2
import numpy as np
from scipy.stats import variation
from openpyxl import Workbook
from openpyxl.utils.cell import get_column_letter

import configs
import najafian.report as report
import najafian.model as model
from configs import FILE_EXTENSION, file_dir
from docker import ensure_dataset_and_output, IN_DOCKER, DATASET_DIR, OUTPUT_DIR
from najafian.find_path import find_file
from najafian.membrane import process_membrane
from najafian.slits import process_slits

# ---- ARGUMENT PARSING ----
parser = argparse.ArgumentParser()

# add argument depending on if in docker
parser.add_argument('input', type=str, help='folder to process')
parser.add_argument('-o', '--output', type=str, help='folder to export to', default=OUTPUT_DIR)
parser.add_argument('-bs', '--batch_size', type=int, help='batch size, depends on system', default=2)
parser.add_argument(
    '--preview', help='show a window preview of segmentations and results', action='store_true')
parser.add_argument(
    '--pskip', help='skip prediction, as in skip creating the segmentation masks and just do post-analysis, to save time', action='store_true')
parser.add_argument(
    '--vskip', help='skip vision/post-processing, use this when you only want to create the segmentation masks', action='store_true'
)
parser.add_argument(
    '--bulk', help='are the results bulked (ie contain glomerulus/subfolders) or process all images together', action='store_true')
parser.add_argument(
    '--use_file_average', help='use the average of file averages and not the average of all FPW measurements (weighted average)', action='store_true')
parser.add_argument(
    '--use_zscore', help='use this to apply the max_slit_zscore which will remove measurements above/below the STDs from the mean', action='store_true')
args = parser.parse_args()


# make sure the specified folders do exist
ensure_dataset_and_output(args.input, args.output)

# now make sure the "constant/globalish" is updated
DATASET_DIR = os.path.abspath(args.input)
OUTPUT_DIR = os.path.abspath(args.output)

# glom only works with bulk
if not args.bulk and args.use_zscore:
    print(f'zscore only works on bulk processes')
    exit(1)

# fix endings to not include last folder option
if str(DATASET_DIR).endswith(os.sep):
    DATASET_DIR = DATASET_DIR[:-1]
if str(OUTPUT_DIR).endswith(os.sep):
    OUTPUT_DIR = OUTPUT_DIR[:-1]
# --- END ARGUMENT PARSING ---


# ---- FILE SPECIFIC CONFIGS ----
# directories to exclude from input processing (usually export only folders)
EXCLUDE_DIRS = ['prediction', 'model', 'gloms', 'std']
SETTINGS = configs.load_project_settings()
CLASSES = SETTINGS['classes']
mark_zscore = 2.8  # any image with X z-score
max_slit_zscore = 5.0  # exclude any slit measurements above a 5.0 z-score
# -- END FILE SPECIFIC CONFIGS --

# ------- LOGGING ------
logger = logging.getLogger('segmentation')
logger.setLevel(logging.INFO)
file_handler = logging.handlers.RotatingFileHandler(
    'analysis.log', maxBytes=2000000, backupCount=5)
file_handler.setFormatter(logging.Formatter(
    '%(asctime)s |%(levelname)-7.7s|:  %(message)s'))
logger.addHandler(file_handler)
console_handler = logging.StreamHandler(sys.stdout)

# try adding colored logs
try:
    import coloredlogs

    formatter = coloredlogs.ColoredFormatter('%(asctime)s |%(levelname)-7.7s|:  %(message)s')
except ImportError:
    logger.warning('could\'t find coloredlogs')
    logger.warning('color support will not be added')

console_handler.setFormatter(formatter)
logger.addHandler(console_handler)

log = logger.info  # default log function
# ------ END LOGGING ----


def ilog(image, title='result', delay=-1):
    """ Simple function to show a numpy array image and rescale to 500x500 for testing/previews

    Args:
        image (np.ndarray): input image (can have 1 or 3 channels)
        title (str, optional): title of window. Defaults to 'result'.
        delay (int, optional): delay of cv2.waitKey. Defaults to -1 (infinite).
    """
    if args.preview:
        cv2.imshow(title, cv2.resize(image, (500, 500)))
        if delay >= 0:
            cv2.waitKey(delay)


def process_post(in_dir, output, bulk=False, gs=None, check_files=None):
    """ Process the layers after the segmentation masks have been created
    please look at membrane.py and slits.py for more details information on processing each layer

    Args:
        in_dir (str): input directory to process
        output (str): the output directory to place files in
        bulk (bool, optional): is this a bulk operation (handles specific stuff like creating the prediction folder). Defaults to False.
        gs (dict, optional): global reporting data (each individual measurement and where to put it on the sheet). Defaults to None.
        check_files (list of str, optional): input file names to compare to. Defaults to None.

    Raises:
        Exception: an error when the mask files don't equal background image length
        Exception: an error when the sorted files have a basename that do not match

    Returns:
        dict: a dictionary of 
    """
    # create the prediction folder for the images if we're not in bulk mode
    # @TODO remove when tested
    # if not bulk:
    #     # note that in_dir is where the masks are located
    #     # then the output folder is just the general output folder (not a specific prediction folder)
    #     in_dir = os.path.join(output, 'prediction')

    #     # add the last sep if it's not there
    #     if in_dir[-1] != os.sep:
    #         in_dir += os.sep

    def sort_unique_files(files):
        """ Gets the unique and sorted list of files with the hidden files removed """
        if files is None:  # ignore this
            return None
        return sorted(configs.unique_everseen(configs.remove_dotfiles(files, sort=True), key=configs.base_split), key=configs.base_split)

    # get all of the processed files (sorted and in order)
    in_files = sort_unique_files(glob.glob(in_dir + '*' + FILE_EXTENSION))
    check_files = sort_unique_files(check_files)

    if check_files is not None and len(in_files) != len(check_files):
        log(f'Input files {str(check_files)}\nMask Files: {str(in_files)}')
        logger.warning(
            f'file count mismatch {len(in_files)} and {len(check_files)}')
        logger.warning(f'for pred folder {in_dir}')
        logger.warning(f'for in files {str(in_files[0])}')
        logger.warning(f'for check files {str(check_files[0])}')
        logger.error(
            'please make sure there are the same number of predicted segmentation masks as there are valid input images! View list above for more info')
        raise Exception('File count mismatch')
    else:
        log('file count match')

    # now let's compare all file names
    if check_files is not None:
        for i, f in zip(in_files, check_files):
            i = configs.base_split(i)
            f = configs.base_split(f)
            if i != f:
                logger.error('mismatched filenames! %s and %s' % (i, f))
                logger.error(
                    'please make sure the file names of segmentation masks and input images are the same')
                raise Exception('Filename mismatch')
    else:
        log('all file checks passed')

    # load some settings
    # settings = configs.load_project_settings()
    log('files: %d' % len(in_files))
    membrane = 0
    membranes = []
    membranes_individual = []
    distances = []
    slit_counts = []
    attachments = []
    attachments_individual = []
    standard_data = []
    id_slits = []
    actual_slits = []
    global_distances = []
    times = []

    # process all of the images
    ind = 0
    for file in in_files:
        log('processing image ' + file)

        # load the post-processed masked image and apply no changes to it
        image = configs.load_image(file, do_fix=False, fix_depth=False)  # load as is
        start = time.time()   # keep track of processing time

        # separate the layers from the stacked image
        membrane_layer = image[0]
        slit_layer = image[1]

        # process the membrane mask layer
        lines, contours, lengths = process_membrane(membrane_layer)

        # process the slit layer using the membrane layer results
        slit_distances, avg_distance, slit_count, ind_attachments, actual_slits = process_slits(slit_layer, contours, lines, lengths, ilog if args.preview else None, ret_more=True)
        
        # this will calculate valid distances across ALL files (for total sum)
        global_distances.extend([float(d) for d in slit_distances if d >= 0])
        
        # capture membrane length of each file
        membrane_length = np.sum(lengths)
        membrane += membrane_length
        membranes.append(membrane_length)

        # get the average (simon's average) later on on a file by file basis
        distances.append(avg_distance)

        # number of slits found in each file
        slit_counts.append(slit_count)

        # this is the % of slit attachment to the membrane (coverage)
        if len(ind_attachments) > 0:
            attachments.append(np.mean(ind_attachments))
        
        # the total (not simon's average) of all membrane lengths will be calculated using this
        membranes_individual.extend([float(length) for length in lengths if length >= 0])

        # i don't think we're currently using this metric but the data is there
        attachments_individual.extend(
            [float(length) for length in ind_attachments if length >= 0])

        # total time elapsed to process this image
        elapsed = time.time() - start
        times.append(elapsed)


        # this seems complicated but essentially if you look at
        # the file mapping dictionary this allows us to easiy
        # map each file to a its respective biopsy and glom
        # so file_mapping[current_file] = {biopsy: ...., glom: ...}
        # then we can keep track of it for our z-score measurements
        # which are stored in the standard_data list
        if gs is not None and check_files is not None:
            _sheet = gs['sheet']
            _mapping = gs['mapping']
            _cur_file = check_files[ind]
            _biopsy = _mapping[_cur_file]['biopsy']
            _glom = _mapping[_cur_file]['glom']

            # OLD global reporting
            # for _distance in slit_distances:
            #    _sheet.append([_biopsy, _glom, _distance])

            standard_data.append({
                'name': file,
                'index': ind,
                'cur_file': _cur_file,
                'biopsy': _biopsy,
                'glom': _glom,
                'measure': [float(m) for m in slit_distances if float(m) >= 0]
            })

        log('finished processing file')
        log('total membrane length %.2f' % membrane)
        log('measurements %d' % len(distances))
        log('time %.2f' % elapsed)

        ind += 1

    # to get the z-scores we have to go back and redo all of the processing
    all_biopsy_gloms = set()
    def b_key(d): return d['biopsy'] + '---' + d['glom']
    for data in standard_data:
        all_biopsy_gloms.add(b_key(data))
        all_biopsy_gloms.add(data['biopsy'] + '---ALL-GLOMS')

    # because we need to process BOTH biopsy and glom there needs to be a new data-structure for-max min std-measurements
    cleaned_up = []

    # process each group of gloms
    for bio_glom in all_biopsy_gloms:
        all_measurements = []
        bio_data = []
        is_all = str(bio_glom).endswith('---ALL-GLOMS')
        for data in standard_data:
            key = (data['biopsy'] + '---ALL-GLOMS') if is_all else b_key(data)
            if key == bio_glom:
                all_measurements.extend(data['measure'])
                bio_data.append(data)

        if len(all_measurements) == 0:
            log('skipping empty glom/biopsy: %s' % bio_glom)
            continue
        else:
            log('PROCESSING: %s' % bio_glom)

        # convert to numpy arr
        all_measure = np.array(all_measurements, dtype=np.float32)
        try:
            mean_all = float(np.average(all_measure))
            std_all = float(np.std(all_measure, dtype=np.float32))
            min_all = float(np.min(all_measure))
            max_all = float(np.max(all_measure))
        except Exception as err:
            log.warning('MALFORMED!', err)
            mean_all = 'N/A'
            std_all = 'N/A'
            min_all = 'N/A'
            max_all = 'N/A'
            continue  # let's skip this as we don't need to restructure a malformed measurement
        log('STD of all measurements: %s (mean: %s)' % (std_all, mean_all))

        # add all files to those specified buckets
        # these buckets (either by each biopsy or each glomerulus)
        # will have z-score calculated for each measurement
        # and if specified they can be removed (only if argument passed)
        def z_calc(x, mean, std): return (
            (float(x) - float(mean)) / float(std))
        for data in bio_data:
            if len(data['measure']) == 0:
                log('no data... skipping z-score test')
                # ws.append([data['biopsy'], data['glom'], data['name']])
                # ws_t.append([data['biopsy'], data['glom'], data['name']])
            else:
                try:
                    av_val = float(np.average(data['measure']))
                except:
                    av_val = 'N/A'
                min_val = min(data['measure'])
                max_val = max(data['measure'])
                if mean_all != 'N/A' and std_all != 'N/A':
                    z_min = z_calc(min_val, mean_all, std_all)
                    z_max = z_calc(max_val, mean_all, std_all)

                    # let's save a copy of the image
                    v_min = None
                    v_max = None
                    if abs(float(z_max)) > max_slit_zscore:
                        v_max = mean_all + (max_slit_zscore * std_all)
                        print(
                            'Skipping measurements for z-score max of %.2f [z-%.2f] (max slit measurement %.2f)' % (v_max, z_max, max_val))
                    if abs(float(z_min)) > max_slit_zscore:
                        v_min = mean_all - (max_slit_zscore * std_all)
                        print(
                            'Skipping measurements for z-score min of %.2f [z-%.2f] (min slit measurement %.2f)' % (v_min, z_min, min_val))

                    # we've passed the threshold
                    if v_min is not None or v_max is not None:
                        print(data['name'], v_min, v_max)
                        cleaned_up.append({
                            'name': data['name'],
                            'cur_file': data['cur_file'],
                            'biopsy': data['biopsy'],
                            'glom': None if is_all else data['glom'],
                            'max': v_max,
                            'min': v_min
                        })

    # now that we know which values that need to be cleaned up
    # read comment above as to why we would do this (to remove extreme)
    # measurements if argument to program is provided
    fixed_distances = []
    if gs is not None and check_files is not None and args.use_zscore:
        log('RESETING global distances as measurements will now use zscore removal')
        global_distances = []   # reset global measurements since we'll use zscore
    for data in standard_data:
        measures = np.array(data['measure'], dtype=np.float32)

        # check for each biopsy/glom z-score min/max requirements
        for check in cleaned_up:
            if data['name'] == check['name'] and data['biopsy'] == check['biopsy'] and (check['glom'] is None or data['glom'] == check['glom']):
                if check['max'] is not None:  # remove max values
                    filtered = measures <= np.float32(check['max'])
                    # print(data['name'], measures, 'max', max(measures))
                    # print('before', measures.shape, 'filter', filtered.shape)
                    before = len(measures)
                    measures = measures[filtered]
                    log('removed %d items for MAX of %.2f' %
                          (before - len(measures), check['max']))
                    # print('after', measures.shape)
                if check['min'] is not None:  # remove min values
                    filtered = measures >= np.float32(check['min'])
                    before = len(measures)
                    measures = measures[filtered]
                    log('removed %d items for MIN of %.2f' %
                          (before - len(measures), check['min']))

        # let's do the fixed distances
        if gs is not None and check_files is not None:
            _sheet = gs['sheet']
            _mapping = gs['mapping']
            _cur_file = data['cur_file']  # check_files[ind]
            _biopsy = _mapping[_cur_file]['biopsy']
            _glom = _mapping[_cur_file]['glom']

            for _distance in measures:
                _sheet.append([_biopsy, _glom, float(_distance)])

        if args.use_zscore:  # only current global distances if z-scores are used
            global_distances.extend([float(m) for m in measures if m >= 0])
        fixed_distances.append(float(np.mean(measures, dtype=np.float32)))

    # replace distances with fixed distances
    # ONLY if argument passed (fixed_distance are distances with high z-scores removed)
    if args.use_zscore:
        distances = fixed_distances

    # calculate the total average
    if args.use_file_average:
        log('using file average!')
        # remove all of the empty distance calculations
        calc_distances = [float(dist) for dist in distances if dist >= 0]
    else:
        log('using global average!')
        # remove all of the empty distance calculations
        calc_distances = [float(dist)
                          for dist in global_distances if dist >= 0]
    
    # optimize arr
    calc_distances = np.array(calc_distances, np.float)

    # we need to handle cases where we can't calculate means
    # or there are invalid values so in the excel/data we will make
    # these numbers negative (ie invalid measurements)
    if len(calc_distances) == 0:
        avg_distance = -1
        std_distance = -1
        cv_distance = -1
        min_max_distance = -1, -1
        avg_slit_count = 0
        std_slit_count = -1
        total_slit_count = -1
        min_max_slit_count = -1, -1
        avg_attachment = 0
        std_attachment = -1
        avg_time = 0
    else:
        avg_distance = float(np.mean(calc_distances, dtype=np.float32))
        std_distance = float(np.std(calc_distances, dtype=np.float32))
        min_max_distance = float(np.amin(calc_distances)), float(
            np.amax(calc_distances))
        cv_distance = float(variation(calc_distances, axis=0))
        avg_slit_count = float(np.mean(slit_counts, dtype=np.float32))
        std_slit_count = float(np.std(slit_counts, dtype=np.float32))
        min_max_slit_count = float(
            np.amin(slit_counts)), float(np.amax(slit_counts))
        total_slit_count = int(np.sum(slit_counts, dtype=np.float32))
        avg_attachment = float(np.mean(attachments, dtype=np.float32))
        std_attachment = float(np.std(attachments, dtype=np.float32))
        avg_time = float(np.mean(times))

    log('end of single image analysis')
    log('total membrane length: %.2f' % membrane)
    log('average slit length: %.2f' % avg_distance)
    log('std slit length: %.2f' % std_distance)
    log('min max slit length: %.2f %.2f' % min_max_distance)
    log('average slit count: %.2f' % avg_slit_count)
    log('std slit count: %.2f' % std_slit_count)
    log('min max slit count %.2f %.2f' % min_max_slit_count)
    log('average attachment %.2f' % avg_attachment)
    log('std attachments %.2f' % std_attachment)
    log('average time %.2f(s)' % avg_time)
    log('saving report to report.json')

    # create the report json
    report_dict = {
        "membrane_length": float(membrane),
        "average_length": avg_distance,
        "std_length": std_distance,
        "cv_length": cv_distance,
        "total_count": total_slit_count,
        "min_max_length": list(min_max_distance),
        "membrane_fraction": float(membrane) / float(total_slit_count),
        "average_count": avg_slit_count,
        "average_attachment": avg_attachment,
        "std_attachment": std_attachment,
        "average_time": avg_time,
        "std_count": std_slit_count,
        "min_max_count": list(min_max_slit_count),
        "file_membrane": [float(length) for length in membranes],
        "file_distances": [float(distance) for distance in distances],
        "file_count": [float(count) for count in slit_counts],
        "file_time": times,
        "global_distances": global_distances,
        "file": [path.basename(f_path) for f_path in in_files],
        "individual_membrane": membranes_individual,
        "individual_attachment": attachments_individual,
        "individual_attachment_percent": [float(attachment / max(membrane, 1.0)) * 100.0 for attachment, membrane in
                                          zip(attachments_individual, membranes_individual)]
    }

    # save the report statistics
    with open(os.path.join(in_dir, 'report.json'), 'w') as r_file:
        r_file.write(json.dumps(report_dict, indent=4))

    # save statistics for histogramming
    with open(os.path.join(in_dir, 'histograms.json'), 'w') as hist_file:
        hist_file.write(json.dumps({
            'std': float(std_distance),
            'cv': float(cv_distance),
            'data': calc_distances.tolist(),
            # 'data': histogram.tolist(),
            # 'edges': np.array(edges).tolist()
        }))

    log('finished saving report')
    return report_dict, standard_data


def process_bulk_post(in_dir, out_dir, prediction_files, prediction_folders, export_ind=None, structure={}):
    log('running bulk post on dataset ' + in_dir)
    prediction_dir = os.path.join(out_dir, 'prediction')

    # add the last sep if it's not there
    if prediction_dir[-1] != os.sep:
        prediction_dir += os.sep

    if not os.path.exists(prediction_dir):
        log('creating prediction folder')
        os.makedirs(prediction_dir)
    else:
        log('prediction folder already exists')

    # create the global report
    wb = Workbook()
    ws = wb.active
    ws.title = 'Bulk Report'
    ws.append(['Report Folder', 'Average File Slit Count', 'Average Time', 'Average FPW',
               'STD FPW', 'Membrane Length', 'Average Attachment', 'STD Attachment', 'Fraction FPW'])

    # do a global measurement book
    gs = wb.create_sheet('Global Measurements')
    gs.append(['Biopsy', 'Glom', 'FPW Measurement (Pixels)'])

    file_index = 0
    # total_files = sum([len(files) for files in prediction_files])
    results = []
    # global_distances = []
    standard_data = []
    file_aggregate_distances = []

    for prediction_file, prediction_folder in zip(prediction_files, prediction_folders):
        log('running post script on folder ' + prediction_folder)
        report_dict, sd = process_post(prediction_folder, out_dir, bulk=True, gs={
            'sheet': gs,
            'mapping': structure
        }, check_files=prediction_file)
        file_index += len(prediction_file)

        # extend the global distances and file aggregates
        # global_distances.extend(report_dict['global_distances'])
        file_aggregate_distances.append([float(d) for d in report_dict['file_distances'] if d > 0])  # append it as a grouped item
        standard_data.extend(sd)

        log('exporting report')
        ws.append([prediction_folder, report_dict['average_count'], report_dict['average_time'], report_dict['average_length'],
                   report_dict['std_length'], report_dict['membrane_length'], report_dict['average_attachment'],
                   report_dict['std_attachment'], report_dict['membrane_fraction']])
        results.append(float(report_dict['average_length']))
        report.compile_report(os.path.join(
            prediction_folder, 'report.xlsx'), report_dict)
        log('done')
        log('done running post script on folder ' + prediction_folder)

    # add the final touches and export the report
    export_file = os.path.join(out_dir, 'bulk_report.xlsx')

    # this is a very useful export as it gives us
    # aggregate results of individual distances that we can use
    # to calculate a running average later on
    with open(os.path.join(out_dir, 'running_average_individual.json'), 'w') as avg_file:
        avg_file.write(json.dumps({
            'data': file_aggregate_distances
        }))

    # attempt to fix the column widths
    # by calculating the max elem size char in that column and setting the
    # overall width to that max char size
    try:
        # calculate the largest item in the column
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column] = max(
                        (dims.get(cell.column, 0), len(str(cell.value))))

        # update each column width
        for col, value in dims.items():
            ws.column_dimensions[get_column_letter(col)].width = value + 1
    except Exception as err:
        logger.error(f'error {str(err)}')

    # group similar gloms (from each biopsy)
    report_std = os.path.join(out_dir, 'std')

    # add the last sep if it's not there
    if report_std[-1] != os.sep:
        report_std += os.sep

    if not os.path.exists(report_std):
        log('creating prediction folder')
        os.makedirs(report_std)
    else:
        log('prediction folder already exists')

    report_std_img = os.path.join(out_dir, 'std-img')

    # add the last sep if it's not there
    if report_std_img[-1] != os.sep:
        report_std_img += os.sep

    if not os.path.exists(report_std_img):
        log('creating prediction folder')
        os.makedirs(report_std_img)
    else:
        log('prediction folder already exists')

    all_biopsy_gloms = set()
    def b_key(d): return d['biopsy'] + '---' + d['glom']
    for data in standard_data:
        all_biopsy_gloms.add(b_key(data))
        all_biopsy_gloms.add(data['biopsy'] + '---ALL-GLOMS')

    """ THIS SECTION OF CODE IS TO GENERATE THE GLOBAL IMAGES WITH THEIR RESPECTIVE Z-SCORES
        * UNCOMMENT IT TO GENERATE THAT REPORT * 
    # create the global report
    wb_t = Workbook()
    ws_t = wb_t.active
    ws_t.title = 'Global Files and STD'
    ws_t.append(['Biopsy', 'Glom', 'File', 'Average FPW', 'Min FPW', 'Max FPW', 'Min Z-score', 'Max Z-score'])
    report_std_global_file = report_std + 'global-std.xlsx'

    # process each group of gloms
    for bio_glom in all_biopsy_gloms:
        report_std_file = report_std + bio_glom + '.xlsx'

        # create the global report
        wb_r = Workbook()
        ws = wb_r.active
        ws.title = 'Files and STD'

        all_measurements = []
        bio_data = []
        is_all = str(bio_glom).endswith('---ALL-GLOMS')
        for data in standard_data:
            key = (data['biopsy'] + '---ALL-GLOMS') if is_all else b_key(data)
            if key == bio_glom:
                all_measurements.extend(data['measure'])
                bio_data.append(data)
        
        if len(all_measurements) == 0:
            print('skipping empty glom/biopsy: %s' % bio_glom)
            continue
        else:
            print('PROCESSING: %s' % bio_glom)

        # convert to numpy arr
        all_measure = np.array(all_measurements, dtype=np.float32)
        try:
            mean_all = float(np.average(all_measure))
            std_all = float(np.std(all_measure, dtype=np.float32))
            min_all = float(np.min(all_measure))
            max_all = float(np.max(all_measure))
        except Exception as err:
            print(err)
            mean_all = 'N/A'
            std_all = 'N/A'
            min_all = 'N/A'
            max_all = 'N/A'
        print('STD of all measurements: %s (mean: %s)' % (std_all, mean_all))

        if is_all:
            ws.append(['Biopsy Average FPW', 'Biopsy STD', 'Biopsy Min', 'Biopsy Max'])
        else:
            ws.append(['Glom Average FPW', 'Glom STD', 'Glom Min', 'Glom Max'])
        ws.append([mean_all, std_all, min_all, max_all])
        ws.append([])
        ws.append(['Biopsy', 'Glom', 'File', 'Average FPW', 'Min FPW', 'Max FPW', 'Min Z-score', 'Max Z-score', 'Type'])

        # add all files to those specified buckets
        z_calc = lambda x, mean, std: ((float(x) - float(mean)) / float(std))
        for data in bio_data:
            if len(data['measure']) == 0:
                ws.append([data['biopsy'], data['glom'], data['name']])
                ws_t.append([data['biopsy'], data['glom'], data['name']])
            else:
                try:
                    av_val = float(np.average(data['measure']))
                except:
                    av_val = 'N/A'
                min_val = min(data['measure'])
                max_val = max(data['measure'])
                if mean_all != 'N/A' and std_all != 'N/A':
                    z_min = z_calc(min_val, mean_all, std_all)
                    z_max = z_calc(max_val, mean_all, std_all)

                    # let's save a copy of the image
                    zval = None
                    if abs(float(z_max)) > mark_zscore:
                        zval = 'max_' + str(z_max) + '_'
                    elif abs(float(z_min)) > mark_zscore:
                        zval = 'min_' + str(z_min) + '_'


                    # we've passed the threshold
                    if zval is not None:
                        print('saving a copy of high z-score image ' )
                        print('z-val: ' + str(zval))

                        report_std_img_folder = report_std_img + bio_glom

                        # add the last sep if it's not there
                        if report_std_img_folder[-1] != os.sep:
                            report_std_img_folder += os.sep

                        if not os.path.exists(report_std_img_folder):
                            log('creating prediction folder')
                            os.makedirs(report_std_img_folder)
                        
                        load_file = data['name']
                        file_name = os.path.splitext(os.path.basename(load_file))[0]  # get the simple base
                        save_name = report_std_img_folder + zval + file_name + '.png' 
                        
                        orig_image_path = os.path.join(base_dir, data['biopsy'], data['glom'])
                        orig_image_path = find_file(orig_image_path, file_name, ['tiff', 'tif', 'png'])
                        print('loading original image of ' + orig_image_path)
                        
                        
                        # load the numpy file
                        orig_image = configs.load_image(orig_image_path) 
                        image = configs.load_image(load_file, fix_depth=False)
                        membrane_layer = image[0]
                        slit_layer = image[1]
                        orig_image = cv2.resize(orig_image, (slit_layer.shape[1], slit_layer.shape[0])).reshape(slit_layer.shape[:2] + (1,))

                        # make an empty image
                        # image = np.zeros((image.shape[1], image.shape[2], 3), np.uint8)
                        # image = configs.mask_image(classes, image, img_shape, orig_image)
                        image = configs.mask_image(classes, image, orig_image.shape, orig_image)
                        

                        # write the alpha mask

                        # make ilog something that saves to a local thing
                        # ilog = lambda img, delay=False: image[np.nonzero(img)] = img[np.nonzero(img)]
                        def ilog(img, delay=False):
                            nonlocal image
                            image[np.nonzero(img)] = img[np.nonzero(img)]

                        # make it a horizontal stack to make it easier to compare
                        img_shape_fix = slit_layer.shape[:2] + (3,)
                        lines, contours, lengths = process_membrane(membrane_layer)
                        fake_data = process_slits(slit_layer, contours, lines, lengths, ilog, ret_more=False)
                        image = np.concatenate((cv2.cvtColor(orig_image, cv2.COLOR_GRAY2BGR).reshape(img_shape_fix), image.reshape(img_shape_fix)), axis=1)
                        cv2.imwrite(save_name, image)
                        print('saved!')
                else:
                    z_min = 'N/A'
                    z_max = 'N/A'
                ws.append([data['biopsy'], data['glom'], data['name'], av_val, min_val, max_val, z_min, z_max, 'Biopsy' if is_all else 'Glom'])
                ws_t.append([data['biopsy'], data['glom'], data['name'], av_val, min_val, max_val, z_min, z_max, 'Biopsy' if is_all else 'Glom'])

        # save the results
        wb_r.save(report_std_file)
    
    # save global results
    wb_t.save(report_std_global_file)
    """

    # save the bulk report
    wb.save(export_file)

    log('exported bulk report to ' + export_file)
    log('finished bulk report')
    log('press the quit button to exit')


def get_files(folder):
    """ Get all files inside a folder by recursively adding to a file list and traversing sub-folders

    Note: prediction and model folders are skipped

    Args:
        folder (str): root folder to scan and continue scanning for valid input background image files

    Returns:
        list of str: a list of file paths 
    """
    files = []
    for f in os.listdir(folder):
        full_path = os.path.join(folder, f)
        if path.isdir(full_path) and f.lower().strip() not in EXCLUDE_DIRS:
            files.extend(get_files(full_path))
        elif configs.is_valid_in_file(full_path)[0]:
            files.append(full_path)
        else:
            log('skipping file/folder %s' % f)
    return files


def get_glom_structure(folder):
    """ Creates a dictionary mapping of the following structure
    {
        "biopsy1": {
            "glom1": ["file1", "file2", ...],
            "glom2": ...
        },
        "biopsy2: {
            "glom1": ...,
            ...
        }
    }

    Finally creating a file mapping where each file has this structure {
        "file1": {
            "biopsy": "biopsy1",
            "glom": "glom1"
        },
        "file2": ...,
        ...
    }

    Args:
        folder (str): the folder with the biopsy/glom structure

    Returns:
        dict: file mapping with the file key to {biopsy: ...., glom: ....} dictionary value mappings
    """
    biopsies = {}
    file_mapping = {}
    for f in os.listdir(folder):
        gloms = {}
        glom_path = os.path.join(folder, f)
        biopsy = f.replace('/', '')

        if path.isdir(glom_path) and f.lower().strip() not in EXCLUDE_DIRS:
            for fg in os.listdir(os.path.join(folder, f)):
                full_path = os.path.join(glom_path, fg)
                if path.isdir(full_path) and fg.lower().strip() not in EXCLUDE_DIRS:
                    files = get_files(full_path)
                    glom = fg.replace('/', '')
                    gloms[glom] = files

                    # add to the global file mapping
                    for full_file in files:
                        file_mapping[full_file] = {
                            'biopsy': biopsy,
                            'glom': glom
                        }
                else:
                    log('skipping glom file/folder %s' % f)
            biopsies[biopsy] = gloms
        else:
            log('skipping biopsy file/folder %s' % f)

    return file_mapping


def get_folders(folder):
    """ Lists the sorted folders in the specified directory """
    folders = []
    for f in os.listdir(folder):
        full_path = os.path.join(folder, f)
        if path.isdir(full_path) and f.lower().strip() not in EXCLUDE_DIRS:
            folders.append((f, full_path))
    folders = sorted(folders, key=lambda x: x[0])
    return [f[1] for f in folders]


def process_folder(in_dir, output):
    """ Processes a single folder in non-bulk mode. Will run predictions if available

    Args:
        in_dir (str): folder to process/analyze
        output (str): the output folder to place the images
    """

    join = os.path.join
    prediction_dir = join(output, 'prediction')

    # add the last sep if it's not there
    if prediction_dir[-1] != os.sep:
        prediction_dir += os.sep

    proc_files = get_files(in_dir)
    if not os.path.exists(prediction_dir):
        log('creating prediction folder')
        os.makedirs(prediction_dir)

    # run the prediction on the folder
    if not args.pskip:
        import najafian.model as model
        for batch, total, files, original, prediction in model.run_prediction_on_folder(proc_files, to_log=True):
            batch_size = len(prediction)
            files = configs.strip_file_names(files)
            for i in range(batch_size):
                configs.save_mask(join(prediction_dir, f'{files[i]}{FILE_EXTENSION}'), prediction[i])

            if batch_size > 0 and len(original) > 0:
                ilog(configs.mask_image(CLASSES, prediction[0], original[0].shape, original[0]), delay=1)
    log('successfully finished the prediction')


def process_bulk_folder(in_dir, out_dir):
    """ Similar to process but this function generates masks for folders in bulk (with sub-folders)

    Args:
        in_dir (str): input folder to generate masks for
        out_dir (str): output folder to place the masks

    Returns:
        tuple: prediction_files (list of str), prediction_folders (list of str)
    """
    folders = get_folders(in_dir)
    join = os.path.join
    file_count = 0
    prediction_files = []
    prediction_folders = []

    for folder in folders:
        rel_path = os.path.relpath(folder, in_dir)
        prediction_dir = join(out_dir, 'prediction', rel_path)

        # add the last sep if it's not there
        if prediction_dir[-1] != os.sep:
            prediction_dir += os.sep

        proc_files = get_files(folder)
        if not os.path.exists(prediction_dir):
            log('creating prediction folder')
            os.makedirs(prediction_dir)
        else:
            log('prediction folder already exists')

        file_count += len(proc_files)
        prediction_files.append(proc_files)
        prediction_folders.append(prediction_dir)

    # similar to function above but applies to all sub-folders
    if not args.pskip:
        import najafian.model as model
        file_index = 0
        for prediction_file, prediction_folder in zip(prediction_files, prediction_folders):
            log('processing bulk folder ' + prediction_folder)
            for batch, total, files, original, prediction in model.run_prediction_on_folder(prediction_file, to_log=True):
                log('bulk progress %d/%d' % (file_index, file_count))
                batch_size = len(prediction)
                files = configs.strip_file_names(files)
                for i in range(batch_size):
                    configs.save_mask(join(prediction_folder, f'{files[i]}{FILE_EXTENSION}'), prediction[i])

                if batch_size > 0 and args.preview:
                    ilog(configs.mask_image(CLASSES, prediction[0], original[0].shape, original[0]), delay=1)

                file_index += len(files)
            log('done processing bulk folder ' + prediction_folder)

    log('successfully finished the prediction')
    return prediction_files, prediction_folders


def process(folder, output, bulk, export_ind=None):
    """ Processes the specified folder either as all images or in bulk and will handle skipping prediction if specified

    Args:
        folder (str): folder to process
        bulk (bool): if True then process each sub-folder 
        export_ind (bool, optional): currently unused (originally for GUI). Defaults to None.
    """
    log('processing folder ' + folder + ' to output ' + output)
    log('is bulk: ' + str(bulk))

    # loads tensorflow if we're going to be predicting
    if not args.pskip:
        log('loading tensorflow...')

        # set basic logging stuff
        from najafian.model import load_model, set_log
        set_log(log)
        log('successfully loaded tensorflow')

        # load the model
        log('loading the model into GPU')
        settings = configs.load_project_settings()
        model_json = os.path.join(file_dir, settings['model'])
        model_weights = os.path.join(file_dir, settings['weights'])
        model_props = os.path.join(file_dir, settings['props'])

        # tells the model module to load the specified model into the GPU
        load_model(model_json, model_weights, model_props, is_layer_model=True, force_reload=False, bs=args.batch_size)
        log('successfully loaded the model into GPU')
    else:
        log('not loading tensorflow... skipping prediction since pskip was specified')

    # are we doing a bulk analysis or a single analysis
    if bulk:
        prediction_files, prediction_folders = process_bulk_folder(folder, output)
        structure = get_glom_structure(folder)

        if args.vskip:
            log('skipping vision processing...')
        else:
            process_bulk_post(folder, output, prediction_files,
                            prediction_folders, export_ind, structure)
    else:
        process_folder(folder, output)
        
        if args.vskip:
            log('skipping vision processing...')
        else:
            process_post(folder, output)
    log('all processing finished')
    log('click quit to exit the program')


if __name__ == '__main__':
    import traceback
    log('Starting analysis. Developed by Najafian Lab')

    try:
        process(folder=DATASET_DIR, output=OUTPUT_DIR, bulk=args.bulk)
    except Exception as err:
        log(f'Failed to process {str(err)}')
        traceback.print_exc()

    log('Complete. Exiting...')
